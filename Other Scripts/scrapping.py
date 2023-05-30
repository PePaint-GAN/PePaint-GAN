import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import os
import warnings
import openpyxl
import pyautogui

warnings.filterwarnings('ignore')


def find_labels(sheet, div, text, column, row):
    span = div.find('span', text=text)
    if span:
        spans = div.find_all('span')
        value = spans[1].text.strip()
        if text=="Código":            
            print(value)
        cell = sheet.cell(row, column)
        cell.value = value

def download_images_and_data(url, total_img, excel_file):
    # Inicializar el controlador
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(url)
    pyautogui.moveTo(600, 530)
    pyautogui.click()
    url = driver.current_url
    
    # Crear un archivo excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Título'
    sheet['B1'] = 'Artista'
    sheet['C1'] = 'Fecha'
    sheet['D1'] = 'Técnica'
    sheet['E1'] = 'Medidas'
    sheet['F1'] = 'Código'
    sheet['G1'] = 'Periodo'
    sheet['H1'] = 'Num'
    row = 2

    # Crear una carpeta para guardar las imágenes
    dir = 'images'
    if not os.path.exists(dir):
        os.makedirs(dir)
    
    # Pasar entre las paginas y realizar las acciones
    resp = requests.get(url, verify=False)
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    for i in range(total_img):
        
        
        # Validar si hay imagen
        labels_img = soup.find_all('img')[1]
        img = labels_img['src']
        id = img.split('/dispatcher/')[1].split('/')[0]
        
        src = 'https://coleccion.mali.pe/internal/media/dispatcher/' + str(id)     
        resp = requests.get(src, stream=True, verify=False)
        if resp.status_code == 200:
        
            
            # Buscar divs de la data
            labels_div = soup.select('div.detailField')
            
            # Guardar los datos en el excel
            for div in labels_div:
                
                title_value = div.find_all('h1')
                if title_value:
                    cell = sheet.cell(row, 1)
                    cell.value = title_value[0].text
                    
                find_labels(sheet, div, 'Artista', 2, row)
                find_labels(sheet, div, 'Fecha', 3, row)
                find_labels(sheet, div, 'Técnica', 4, row)
                find_labels(sheet, div, 'Medidas', 5, row)
                find_labels(sheet, div, 'Código', 6, row)
            
            # Guardar imagen        
            # Obtener la extensión de archivo correcta
            content_type = resp.headers.get('content-type')
            if 'image/jpeg' in content_type:
                extension = 'jpg'
            elif 'image/png' in content_type:
                extension = 'png'
            

            # Generar un nombre de archivo único
            file_name = os.path.join(dir, f'Republicano-{row-1}.{extension}')
            
            # Guardar la imagen en disco
            with open(file_name, 'wb') as file:
                for chunk in resp.iter_content(1024):
                    file.write(chunk)
                
        
            # Guardar la id
            cell = sheet.cell(row, 7)
            cell.value = "Republicano"
            
            cell = sheet.cell(row, 8)
            cell.value = row-1
            
            
            
            row += 1
        
        else:
            print("No hay imagen")
        
        # Pasar a la siguiente pagina        
        pyautogui.moveTo(1650, 250)
        pyautogui.click()
        url = driver.current_url
        
        resp = requests.get(url, verify=False)
        soup = BeautifulSoup(resp.text, 'html.parser')
            
    workbook.save(excel_file)

    print(f'Archivo Excel guardado: {excel_file}')
    
    driver.quit()


url = 'https://coleccion.mali.pe/objects/images?filter=classifications%3APintura%3Bcollections%3A3.%20Republicano%20%20%28S.%20XIX%20y%20XX%29'
total_img = 1695
excel_file = 'PaintGAN_Republicano.xlsx'

download_images_and_data(url, total_img, excel_file)